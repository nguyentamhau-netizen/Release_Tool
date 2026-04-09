from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from release_note_tool.core import find_matching_sheets, generate_test_result


def create_sample_release_note(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sprint 6"
    headers = ["No", "Sprint", "Service", "Module", "Link", "US ID", "Type", "US Name"]
    for index, header in enumerate(headers, start=1):
        worksheet.cell(row=3, column=index).value = header

    worksheet["A4"] = 1
    worksheet["B4"] = "Sprint 6"
    worksheet["C4"] = "Seller Center"
    worksheet["D4"] = "Shop"
    worksheet["E4"] = "https://example.com/ticket/7738"
    worksheet["F4"] = "7738"
    worksheet["G4"] = "User Story"
    worksheet["H4"] = "[Seller Center][FE] UI enhancement"

    workbook.save(path)


def run_smoke_test() -> None:
    with TemporaryDirectory() as temp_dir:
        temp_root = Path(temp_dir)
        input_path = temp_root / "release-note.xlsx"
        output_dir = temp_root / "generated"
        create_sample_release_note(input_path)

        sheets = find_matching_sheets(input_path)
        assert sheets == ["Sprint 6"]

        output_path = generate_test_result(
            input_path=input_path,
            sheet_name="Sprint 6",
            release_type="UAT",
            request_date="09-04-2026",
            output_dir=output_dir,
        )

        workbook = load_workbook(output_path)
        worksheet = workbook["Summary"]
        assert worksheet["A8"].value == "1"
        assert worksheet["F8"].value == "7738"
        assert worksheet["I8"].value == "[Seller Center][FE] UI enhancement"


if __name__ == "__main__":
    run_smoke_test()
    print("Smoke test passed.")
