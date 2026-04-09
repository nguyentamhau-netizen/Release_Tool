from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from .taiga import TaigaClient, TaigaConfig


REQUIRED_SOURCE_HEADERS = (
    "No",
    "Sprint",
    "Service",
    "Module",
    "Link",
    "Type",
    "US Name",
)
ID_HEADERS = (
    "US ID",
    "IS ID",
    "Issue ID",
)
OUTPUT_HEADERS = (
    "No",
    "Sprint",
    "Service",
    "Module",
    "Link",
    "US ID",
    "Status",
    "QC PIC",
    "US Name",
)
HEADER_ALIASES = {
    "us name": "US Name",
    "us_name": "US Name",
    "us id": "US ID",
    "us_id": "US ID",
    "is id": "IS ID",
    "is_id": "IS ID",
    "issue id": "Issue ID",
    "issue_id": "Issue ID",
    "modul": "Module",
    "issue/us": "Type",
}


@dataclass(frozen=True)
class ReleaseRow:
    no: str
    sprint: str
    service: str
    module: str
    link: str
    us_id: str
    item_type: str
    us_name: str


@dataclass(frozen=True)
class TaigaLogEntry:
    row_no: str
    item_type: str
    ref_id: str
    status: str
    message: str


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().replace("\n", " ")
    text = re.sub(r"\s+", " ", text)
    alias = HEADER_ALIASES.get(text.casefold())
    if alias:
        return alias.casefold()
    return text.casefold()


def stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def sanitize_filename(value: str) -> str:
    return re.sub(r'[<>:"/\\\\|?*]+', "-", value).strip()


def build_output_filename(release_type: str, request_date: str) -> str:
    return sanitize_filename(
        f"KD - Test Result - {release_type} Request {request_date}.xlsx"
    )


def find_matching_sheets(workbook_path: Path) -> List[str]:
    workbook = load_workbook(workbook_path, data_only=False)
    matches: List[str] = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        try:
            find_header_map(worksheet)
        except ValueError:
            continue
        matches.append(sheet_name)
    return matches


def find_header_map(worksheet: Worksheet) -> tuple[int, Dict[str, int]]:
    best_row = -1
    best_map: Dict[str, int] = {}
    normalized_required = {
        header.casefold(): header
        for header in (*REQUIRED_SOURCE_HEADERS, *ID_HEADERS)
    }

    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, values_only=True),
        start=1,
    ):
        current: Dict[str, int] = {}
        for col_index, value in enumerate(row, start=1):
            normalized = normalize_header(value)
            if normalized in normalized_required:
                current[normalized_required[normalized]] = col_index
        if len(current) > len(best_map):
            best_row = row_index
            best_map = current
        if _has_required_headers(best_map):
            return best_row, best_map

    missing = [header for header in REQUIRED_SOURCE_HEADERS if header not in best_map]
    if not any(header in best_map for header in ID_HEADERS):
        missing.append("one of: US ID / IS ID / Issue ID")
    raise ValueError(
        f"Sheet '{worksheet.title}' is missing required headers: {', '.join(missing)}"
    )


def read_release_rows(workbook_path: Path, sheet_name: str) -> List[ReleaseRow]:
    workbook = load_workbook(workbook_path, data_only=False)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")

    worksheet = workbook[sheet_name]
    header_row, header_map = find_header_map(worksheet)
    rows: List[ReleaseRow] = []
    empty_streak = 0

    for row_index in range(header_row + 1, worksheet.max_row + 1):
        row = {
            header: stringify(worksheet.cell(row=row_index, column=column).value)
            for header, column in header_map.items()
        }
        if not row["No"]:
            empty_streak += 1
            if empty_streak >= 3:
                break
            continue

        empty_streak = 0
        item_type = row.get("Type") or "User Story"
        item_id = pick_item_id(item_type, row)
        if not item_id:
            continue

        rows.append(
            ReleaseRow(
                no=row["No"],
                sprint=row["Sprint"],
                service=row["Service"],
                module=row["Module"],
                link=row["Link"],
                us_id=item_id,
                item_type=item_type,
                us_name=row["US Name"],
            )
        )

    if not rows:
        raise ValueError(f"No data rows found in sheet '{sheet_name}'.")

    return rows


def create_test_result_workbook(
    rows: Sequence[ReleaseRow],
    release_type: str,
    request_date: str,
) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Summary"
    layout = build_summary_layout(worksheet, release_type, request_date)
    write_data_rows(worksheet, rows, data_start_row=layout["data_start_row"])
    if layout["has_summary"]:
        write_status_summary_formulas(
            worksheet,
            row_count=len(rows),
            data_start_row=layout["data_start_row"],
        )
    return workbook


def build_summary_layout(
    worksheet: Worksheet,
    release_type: str,
    request_date: str,
) -> Dict[str, int | bool]:
    green_fill = PatternFill("solid", fgColor="93C47D")
    dark_fill = PatternFill("solid", fgColor="D9EAD3")
    white_font = Font(bold=True, color="FFFFFF")
    header_font = Font(bold=True)
    thin_side = Side(style="thin", color="000000")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    normalized_release_type = release_type.strip().upper()
    has_summary = normalized_release_type == "PROD"
    header_row = 7 if has_summary else 1

    if has_summary:
        summary_labels = [
            ("B2", "Release Type"),
            ("B3", "Generated Date"),
            ("B4", "Passed Dev"),
            ("B5", "Passed UAT"),
            ("E4", "Passed Pro"),
            ("E5", "On Prod"),
        ]
        values = {
            "C2": release_type,
            "C3": request_date,
        }

        for cell_ref, text in summary_labels:
            worksheet[cell_ref] = text
            worksheet[cell_ref].font = header_font
            worksheet[cell_ref].fill = dark_fill
            worksheet[cell_ref].border = border
            worksheet[cell_ref].alignment = Alignment(horizontal="center")

        for cell_ref, value in values.items():
            worksheet[cell_ref] = value
            worksheet[cell_ref].border = border
            worksheet[cell_ref].alignment = Alignment(horizontal="center")

        for cell_ref in ("C4", "C5", "F4", "F5"):
            worksheet[cell_ref].fill = PatternFill("solid", fgColor="FFFFFF")
            worksheet[cell_ref].border = border
            worksheet[cell_ref].alignment = Alignment(horizontal="center")

    for index, header in enumerate(OUTPUT_HEADERS, start=1):
        cell = worksheet.cell(row=header_row, column=index)
        cell.value = header
        cell.fill = green_fill
        cell.font = white_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    data_start_row = header_row + 1
    worksheet.freeze_panes = f"A{data_start_row}"
    worksheet.auto_filter.ref = f"A{header_row}:I{header_row}"

    widths = {
        "A": 8,
        "B": 14,
        "C": 20,
        "D": 18,
        "E": 32,
        "F": 12,
        "G": 18,
        "H": 18,
        "I": 60,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    return {
        "has_summary": has_summary,
        "header_row": header_row,
        "data_start_row": data_start_row,
    }


def write_data_rows(
    worksheet: Worksheet,
    rows: Sequence[ReleaseRow],
    data_start_row: int,
) -> None:
    thin_side = Side(style="thin", color="000000")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for offset, row in enumerate(rows):
        row_index = data_start_row + offset
        values = (
            row.no,
            row.sprint,
            row.service,
            row.module,
            row.link,
            row.us_id,
            "",
            "",
            row.us_name,
        )
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical="center")
        if row.link:
            worksheet.cell(row=row_index, column=5).hyperlink = row.link
            worksheet.cell(row=row_index, column=5).style = "Hyperlink"


def write_status_summary_formulas(
    worksheet: Worksheet,
    row_count: int,
    data_start_row: int,
) -> None:
    last_row = max(data_start_row, data_start_row + row_count - 1)
    status_range = f"$G${data_start_row}:$G${last_row}"
    worksheet["C4"] = f'=COUNTIF({status_range},"*PASSED DEV*")'
    worksheet["C5"] = f'=COUNTIF({status_range},"*PASSED UAT*")'
    worksheet["F4"] = f'=COUNTIF({status_range},"*PASSED PRO*")'
    worksheet["F5"] = f'=COUNTIF({status_range},"*ON PROD*")'


def build_taiga_enrichment(
    rows: Sequence[ReleaseRow],
    taiga_config_path: Optional[Path],
) -> tuple[Dict[str, Dict[str, str]], List[TaigaLogEntry]]:
    if not taiga_config_path or not taiga_config_path.exists():
        return {}, [
            TaigaLogEntry(
                row_no="-",
                item_type="-",
                ref_id="-",
                status="INFO",
                message="taiga.local.json not found. Status and QC PIC were left blank.",
            )
        ]

    client = TaigaClient(TaigaConfig.from_path(taiga_config_path))
    enrichment: Dict[str, Dict[str, str]] = {}
    logs: List[TaigaLogEntry] = []
    for row in rows:
        ref = row.us_id.strip()
        if not ref:
            logs.append(
                TaigaLogEntry(
                    row_no=row.no,
                    item_type=row.item_type,
                    ref_id=ref,
                    status="SKIP",
                    message="Missing ref id after normalization.",
                )
            )
            continue
        try:
            result = client.enrich(row.item_type, ref)
            enrichment[ref] = result
            logs.append(
                TaigaLogEntry(
                    row_no=row.no,
                    item_type=row.item_type,
                    ref_id=ref,
                    status="OK",
                    message=(
                        f"source='{result.get('_source', '')}', "
                        f"Status='{result.get('Status', '')}', "
                        f"raw PIC='{result.get('_raw_qc_pic', '')}', "
                        f"filtered QC PIC='{result.get('QC PIC', '')}'"
                    ),
                )
            )
        except Exception as exc:
            enrichment[ref] = {"Status": "", "QC PIC": ""}
            logs.append(
                TaigaLogEntry(
                    row_no=row.no,
                    item_type=row.item_type,
                    ref_id=ref,
                    status="ERROR",
                    message=str(exc),
                )
            )
    return enrichment, logs


def apply_taiga_enrichment(
    worksheet: Worksheet,
    rows: Sequence[ReleaseRow],
    taiga_map: Dict[str, Dict[str, str]],
    data_start_row: int = 8,
) -> None:
    for offset, row in enumerate(rows):
        row_index = data_start_row + offset
        values = taiga_map.get(row.us_id, {})
        worksheet.cell(row=row_index, column=7).value = values.get("Status", "")
        worksheet.cell(row=row_index, column=8).value = values.get("QC PIC", "")


def generate_test_result(
    input_path: Path,
    sheet_name: str,
    release_type: str,
    request_date: str,
    output_dir: Path,
    taiga_config_path: Optional[Path] = None,
) -> Path:
    rows = read_release_rows(input_path, sheet_name)
    workbook = create_test_result_workbook(rows, release_type, request_date)
    data_start_row = 8 if release_type.strip().upper() == "PROD" else 2
    taiga_map, taiga_logs = build_taiga_enrichment(rows, taiga_config_path)
    apply_taiga_enrichment(
        workbook["Summary"],
        rows,
        taiga_map,
        data_start_row=data_start_row,
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / build_output_filename(release_type, request_date)
    workbook.save(output_path)
    write_taiga_log(output_path, taiga_logs)
    return output_path


def today_text() -> str:
    return date.today().strftime("%d-%m-%Y")


def _has_required_headers(header_map: Dict[str, int]) -> bool:
    return all(header in header_map for header in REQUIRED_SOURCE_HEADERS) and any(
        header in header_map for header in ID_HEADERS
    )


def pick_item_id(item_type: str, row: Dict[str, str]) -> str:
    normalized_type = item_type.strip().casefold()
    if normalized_type == "issue":
        candidates = ("IS ID", "Issue ID", "US ID")
    else:
        candidates = ("US ID", "IS ID", "Issue ID")

    for header in candidates:
        value = row.get(header, "").strip()
        if value:
            return normalize_ref_id(value)
    return ""


def normalize_ref_id(value: str) -> str:
    text = value.strip()
    match = re.search(r"(\d+)", text)
    if match:
        return match.group(1)
    return text


def write_taiga_log(output_path: Path, logs: Sequence[TaigaLogEntry]) -> None:
    log_path = output_path.with_name(output_path.stem + " - taiga-log.txt")
    lines = [
        "Taiga Lookup Log",
        f"Output File: {output_path.name}",
        "",
    ]
    for log in logs:
        lines.append(
            f"[{log.status}] Row {log.row_no} | {log.item_type} | Ref {log.ref_id} | {log.message}"
        )
    log_path.write_text("\n".join(lines), encoding="utf-8")
